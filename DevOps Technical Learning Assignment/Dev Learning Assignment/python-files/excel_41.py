import openpyxl as xl  # this is an alias for this package
from openpyxl.chart import BarChart, Reference  # importing chart

# in this package we have a module called chart. from this module we are importing two classes which are barchart and refernce. case sensitive.


def process_workbook(filename):
    wb = xl.load_workbook(
        filename
    )  # this will load our excel workbook and return a workbook object.

    # if u open workbook there is only one sheet.
    sheet = wb["Sheet1"]  # case sensitive.

    # cell = sheet["a1"]  # to access a certain cell
    # cell = sheet.cell(1, 1)# same thing

    # print(cell.value) #pints out value from a1

    # print(sheet.max_row)  # shows amount of rows.
    # make a for loop that would generate the number 1 and 4.

    for row in range(2, sheet.max_row + 1):  # to include the 4th, add one to it.
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        correct_price_cell = sheet.cell(row, 4)
        correct_price_cell.value = corrected_price

    values = Reference(
        sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4
    )  # selecting cells from row 2 to 4. limintg range of cells we are slecting to the 4th cloum.

    chart = BarChart()
    chart.add_data(values)

    sheet.add_chart(chart, "e2")
    # wb.save("transactions2.xlsx")
    wb.save(filename)
    # can overwrite old prices or add new colum
